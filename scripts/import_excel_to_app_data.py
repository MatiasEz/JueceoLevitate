#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import datetime
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import warnings as py_warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

py_warnings.filterwarnings(
    "ignore",
    message="DrawingML support is incomplete*",
    category=UserWarning,
    module="openpyxl.reader.drawings",
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "JueceoCoreografias" / "Resources" / "app_data.json"
DEFAULT_XLSX = ROOT / "JueceoCoreografias" / "Resources" / "Bloque2.xlsx"
DEFAULT_TEMPLATE_XLSX = DEFAULT_XLSX
DEFAULT_ENV = ROOT / ".env"
DEFAULT_JUDGE_PROFILES = [
    {"name": "ALEX", "hero_image_name": "JudgeHeroAlex"},
    {"name": "ANGELA", "hero_image_name": "JudgeHeroAngela"},
    {"name": "DANIEL", "hero_image_name": "JudgeHeroDaniel"},
    {"name": "DAVE", "hero_image_name": ""},
    {"name": "EVA", "hero_image_name": ""},
    {"name": "VLADIMIR", "hero_image_name": "JudgeHeroVladimir"},
    {"name": "YOLI", "hero_image_name": "JudgeHeroYoli"},
    {"name": "ATI", "hero_image_name": ""},
]
DEFAULT_JUDGES = [profile["name"] for profile in DEFAULT_JUDGE_PROFILES]

REQUIRED_BLOCK_COLUMNS = {
    "name": "COREOGRAFIA",
    "academy": "ACADEMIA",
    "division": "DIVISION",
    "genre": "GENERO/SUBGENERO",
    "category": "CATEGORIA",
}

TEMPLATE_ALIASES = {
    "ARO": "DANZA AEREA",
    "TELA": "DANZA AEREA",
    "URBANOS": "HIP HOP",
    "OPEN": "CONTEMPORANEO",
}


@dataclass
class ImportResult:
    app_data: dict[str, Any]
    warnings: list[str]
    errors: list[str]


def normalized(value: Any) -> str:
    text = str(value or "").strip().upper()
    return "".join(
        character
        for character in unicodedata.normalize("NFD", text)
        if unicodedata.category(character) != "Mn"
    )


def stable_id(value: Any) -> str:
    base = normalized(value).lower()
    base = re.sub(r"[^a-z0-9]+", "-", base)
    base = re.sub(r"-+", "-", base).strip("-")
    return base or "sin-dato"


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def row_text(row: list[Any], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return clean_text(row[index])


def compact_number_text(value: Any) -> str:
    text = clean_text(value).replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def routine_id_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return clean_text(value).replace(".0", "")


def cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime.time):
        if value.second or value.microsecond:
            return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime.datetime):
        if value.year == 1900:
            return str(value.day)
        return value.isoformat()
    return value


def sheet_rows(sheet) -> list[list[Any]]:
    return [
        [cell_value(sheet.cell(row_index, column_index).value) for column_index in range(1, sheet.max_column + 1)]
        for row_index in range(1, sheet.max_row + 1)
    ]


def inferred_duration(formula_rows: list[list[Any]], row_number: int, time_index: int | None) -> str:
    if time_index is None:
        return ""
    next_row_index = row_number
    if next_row_index >= len(formula_rows) or time_index >= len(formula_rows[next_row_index]):
        return ""
    next_time_formula = formula_rows[next_row_index][time_index]
    if not isinstance(next_time_formula, str):
        return ""

    match = re.match(
        r"^=\s*\$?[A-Z]+\$?(\d+)\s*\+\s*([0-9]+(?:[.,][0-9]+)?)\s*/\s*1440\s*$",
        next_time_formula,
        flags=re.IGNORECASE,
    )
    if not match or int(match.group(1)) != row_number:
        return ""
    return compact_number_text(match.group(2))


def load_env(path: Path = DEFAULT_ENV) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        clean = line.strip()
        if not clean or clean.startswith("#") or "=" not in clean:
            continue
        key, value = clean.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def parse_block(sheet, formula_sheet=None) -> tuple[dict[str, Any] | None, list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    rows = sheet_rows(sheet)
    formula_rows = sheet_rows(formula_sheet) if formula_sheet is not None else rows
    header_index = None
    for index, row in enumerate(rows):
        headers = [normalized(value) for value in row]
        if "COREOGRAFIA" in headers and "ACADEMIA" in headers:
            header_index = index
            break
    if header_index is None:
        return None, warnings, errors

    column_map: dict[str, int] = {}
    for index, value in enumerate(rows[header_index]):
        key = normalized(value)
        if "COREOGRAFIA" in key:
            column_map["name"] = index
        if "ACADEMIA" in key:
            column_map["academy"] = index
        if "DIVISION" in key:
            column_map["division"] = index
        if "GENERO" in key or "MODALIDAD" in key:
            column_map["genre"] = index
        if "NIVEL" in key:
            column_map["level"] = index
        if "CATEGORIA" in key:
            column_map["category"] = index
        if "COREOGRAFO" in key:
            column_map["choreographer"] = index
        if "PARTICIPANTE" in key:
            column_map["participant"] = index
        if "ESTADO" in key:
            column_map["state"] = index
        if "HORARIO" in key:
            column_map["time"] = index
        if "DURACION" in key:
            column_map["duration"] = index

    missing = [label for key, label in REQUIRED_BLOCK_COLUMNS.items() if key not in column_map]
    if missing:
        warnings.append(f"{sheet.title}: parece una hoja auxiliar, se omite como bloque; faltan {', '.join(missing)}.")
        return None, warnings, errors

    block_id = stable_id(sheet.title)
    routines: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not row or not str(row[0]).strip():
            continue
        name = row_text(row, column_map["name"])
        if not name:
            warnings.append(f"{sheet.title} fila {row_number}: coreografia vacia, se omite.")
            continue
        routine_id = routine_id_text(row[0])
        if not routine_id:
            warnings.append(f"{sheet.title} fila {row_number}: numero de rutina vacio, se omite.")
            continue
        duration = row_text(row, column_map.get("duration"))
        if not duration:
            duration = inferred_duration(formula_rows, row_number, column_map.get("time"))
        routines.append(
            {
                "id": routine_id,
                "blockID": block_id,
                "block": sheet.title,
                "name": name.title(),
                "academy": row_text(row, column_map.get("academy")),
                "division": row_text(row, column_map.get("division")),
                "genre": row_text(row, column_map.get("genre")),
                "level": row_text(row, column_map.get("level")),
                "category": row_text(row, column_map.get("category")),
                "choreographer": row_text(row, column_map.get("choreographer")),
                "participant": row_text(row, column_map.get("participant")),
                "state": row_text(row, column_map.get("state")),
                "time": row_text(row, column_map.get("time")),
                "duration": duration,
            }
        )

    if not routines:
        return None, warnings, errors

    try:
        sort_order = int(str(routines[0]["id"]).strip())
    except (IndexError, ValueError):
        sort_order = 0
    title = " · ".join(str(value) for row in rows[:header_index] for value in row if value)
    return {
        "blockID": block_id,
        "name": sheet.title,
        "title": title,
        "sortOrder": sort_order,
        "isActive": False,
        "routines": routines,
    }, warnings, errors


def parse_template(sheet) -> tuple[dict[str, Any] | None, list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    sheet_key = normalized(sheet.title)
    if sheet_key.startswith("JUECEO") or sheet_key.startswith("COPIA DE JUECEO") or sheet_key in {
        "CALIFICACIONES",
        "DICTAMEN FINAL",
    }:
        return None, warnings, errors

    rows = sheet_rows(sheet)
    title = next(
        (str(value) for row in rows for value in row if "HOJA DE JUECEO" in normalized(value)),
        None,
    )
    template_markers = {"TECNICA", "EJECUCION COREOGRAFICA", "INTERPRETACION", "ARTISTICO"}
    has_template_marker = any(
        normalized(value) in template_markers
        for row in rows
        for value in row
        if value
    )

    section = "General"
    criteria: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=1):
        number_value = row[1] if len(row) > 1 else ""
        label = str(row[2] if len(row) > 2 else "").strip()
        max_value = row[3] if len(row) > 3 else ""
        try:
            number = float(str(number_value).replace(",", "."))
        except ValueError:
            number = 0
        try:
            maximum = float(str(max_value).replace(",", "."))
        except ValueError:
            maximum = 0

        if number_value and not number and normalized(number_value) not in ["-", "JUEZ", "FEEDBACK"]:
            section = str(number_value).strip()
        if number > 0 and label:
            if maximum <= 0:
                errors.append(f"{sheet.title} fila {row_number}: criterio '{label}' sin puntaje maximo.")
                continue
            criteria.append(
                {"id": int(number), "section": section, "label": label, "maxScore": maximum}
            )

    if not criteria:
        if title:
            errors.append(f"{sheet.title}: hoja de jueceo sin criterios validos.")
        return None, warnings, errors
    if not title:
        if not has_template_marker:
            return None, warnings, errors
        title = f"HOJA DE JUECEO {sheet.title}"

    return {
        "genre": sheet.title,
        "title": title,
        "maxScore": sum(item["maxScore"] for item in criteria),
        "criteria": criteria,
    }, warnings, errors


def parse_judges(workbook) -> tuple[list[str], list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    judges: list[str] = []
    seen: set[str] = set()
    duplicate_count = 0
    if "CALIFICACIONES" in workbook.sheetnames:
        sheet = workbook["CALIFICACIONES"]
        for row_index in range(1, sheet.max_row + 1):
            value = sheet.cell(row_index, 3).value
            if not value or normalized(value) == "JUEZ" or str(value).startswith("="):
                continue
            judge = str(value).strip().upper()
            key = normalized(judge)
            if not key:
                continue
            if key in seen:
                duplicate_count += 1
                continue
            seen.add(key)
            judges.append(judge)
    if not judges:
        judges = DEFAULT_JUDGES.copy()
        warnings.append("No se encontraron jueces en CALIFICACIONES; se usan jueces default de Levitate.")
    elif duplicate_count:
        warnings.append(f"CALIFICACIONES: {duplicate_count} jueces duplicados omitidos.")
    return judges, warnings, errors


def with_required_people(judges: list[str]) -> list[str]:
    result = DEFAULT_JUDGES.copy()
    known = {stable_id(judge) for judge in result}
    for judge in judges:
        judge_id = stable_id(judge)
        if judge_id not in known:
            result.append(judge)
            known.add(judge_id)
    return result


def role_for_person(name: str) -> str:
    return "admin" if stable_id(name) == "ati" else "judge"


def hero_image_for_person(name: str) -> str:
    judge_id = stable_id(name)
    for profile in DEFAULT_JUDGE_PROFILES:
        if stable_id(profile["name"]) == judge_id:
            return profile["hero_image_name"]
    return ""


def load_templates(source: Path) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    if not source.exists():
        return [], [f"No se encontro fuente de plantillas: {source}."], []

    workbook = load_workbook(source, data_only=True)
    warnings: list[str] = []
    errors: list[str] = []
    templates: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        template, template_warnings, template_errors = parse_template(sheet)
        warnings.extend(template_warnings)
        errors.extend(template_errors)
        if template:
            templates.append(template)
    return templates, warnings, errors


def alias_base_for_genre(genre: str) -> str | None:
    genre_key = normalized(genre)
    if genre_key.startswith("OPEN:"):
        return "DANZA AEREA"
    return TEMPLATE_ALIASES.get(genre_key)


def aliased_template(base_template: dict[str, Any], genre: str) -> dict[str, Any]:
    template = copy.deepcopy(base_template)
    template["genre"] = genre
    template["title"] = f"HOJA DE JUECEO {genre}"
    return template


def ensure_templates_for_routines(
    templates: list[dict[str, Any]],
    routines: list[dict[str, Any]],
) -> tuple[list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    templates_by_genre = {normalized(template["genre"]): template for template in templates}
    for genre in sorted({routine["genre"] for routine in routines if routine.get("genre")}, key=normalized):
        genre_key = normalized(genre)
        if genre_key in templates_by_genre:
            continue

        base_key = alias_base_for_genre(genre)
        base_template = templates_by_genre.get(base_key or "")
        if base_template:
            templates.append(aliased_template(base_template, genre))
            templates_by_genre[genre_key] = templates[-1]
            warnings.append(f"{genre}: se usa la plantilla base {base_template['genre']}.")
            continue

        errors.append(f"Genero sin hoja de jueceo: {genre}.")
    return warnings, errors


def parse_workbook(source: Path, template_source: Path | None = DEFAULT_TEMPLATE_XLSX) -> ImportResult:
    workbook = load_workbook(source, data_only=True)
    formula_workbook = load_workbook(source, data_only=False)
    warnings: list[str] = []
    errors: list[str] = []
    blocks: list[dict[str, Any]] = []
    templates: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        formula_sheet = formula_workbook[sheet.title] if sheet.title in formula_workbook.sheetnames else None
        block, block_warnings, block_errors = parse_block(sheet, formula_sheet)
        warnings.extend(block_warnings)
        errors.extend(block_errors)
        if block:
            blocks.append(block)

        template, template_warnings, template_errors = parse_template(sheet)
        warnings.extend(template_warnings)
        errors.extend(template_errors)
        if template:
            templates.append(template)

    routines = [routine for block in blocks for routine in block["routines"]]
    if routines and not templates and template_source:
        fallback_templates, fallback_warnings, fallback_errors = load_templates(template_source)
        warnings.extend(fallback_warnings)
        errors.extend(fallback_errors)
        if fallback_templates:
            templates = fallback_templates
            warnings.append(f"Se usan plantillas base desde {template_source.name}.")

    judges, judge_warnings, judge_errors = parse_judges(workbook)
    judges = with_required_people(judges)
    warnings.extend(judge_warnings)
    errors.extend(judge_errors)

    routine_ids: set[str] = set()
    for routine in routines:
        if routine["id"] in routine_ids:
            errors.append(f"Rutina duplicada: #{routine['id']}.")
        routine_ids.add(routine["id"])

    template_warnings, template_errors = ensure_templates_for_routines(templates, routines)
    warnings.extend(template_warnings)
    errors.extend(template_errors)

    app_data = {
        "sourceName": source.name,
        "blocks": blocks,
        "routines": routines,
        "templates": templates,
        "judges": judges,
        "judgeProfiles": [
            {
                "judgeID": stable_id(judge),
                "name": judge,
                "role": role_for_person(judge),
                "heroImageName": hero_image_for_person(judge),
            }
            for judge in judges
        ],
    }
    return ImportResult(app_data=app_data, warnings=warnings, errors=errors)


def write_json(app_data: dict[str, Any], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(app_data, ensure_ascii=False, indent=2), encoding="utf-8")


def supabase_request(
    method: str,
    path: str,
    payload: Any | None = None,
    *,
    prefer: str | None = None,
) -> Any:
    base_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        or os.environ.get("SUPABASE_SECRET_KEY")
    )
    if not base_url or not key:
        raise RuntimeError("Faltan SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY/SUPABASE_SECRET_KEY.")

    data = None if payload is None else json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(f"{base_url}/rest/v1/{path.lstrip('/')}", data=data, method=method)
    request.add_header("apikey", key)
    request.add_header("Authorization", f"Bearer {key}")
    request.add_header("Content-Type", "application/json")
    if prefer:
        request.add_header("Prefer", prefer)

    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else None
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8")
        raise RuntimeError(f"Supabase {method} {path} fallo: {error.code} {detail}") from error


def chunked(items: list[dict[str, Any]], size: int = 500):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def upsert_rows(table: str, rows: list[dict[str, Any]], conflict: str) -> None:
    if not rows:
        return
    query = urllib.parse.urlencode({"on_conflict": conflict})
    for chunk in chunked(rows):
        try:
            supabase_request(
                "POST",
                f"{table}?{query}",
                chunk,
                prefer="resolution=merge-duplicates,return=minimal",
            )
        except RuntimeError as error:
            fallback_columns = {
                "judges": ["role", "hero_image_name"],
                "routines": ["participant"],
            }.get(table, [])
            missing_columns = [column for column in fallback_columns if column in str(error)]
            if not missing_columns:
                raise
            fallback_chunk = [
                {key: value for key, value in row.items() if key not in missing_columns}
                for row in chunk
            ]
            supabase_request(
                "POST",
                f"{table}?{query}",
                fallback_chunk,
                prefer="resolution=merge-duplicates,return=minimal",
            )


def delete_event_rows(event_id: str) -> None:
    encoded = urllib.parse.quote(event_id)
    for table in ["penalties", "feedback", "scores", "criteria", "criteria_templates", "judges", "routines", "blocks"]:
        supabase_request("DELETE", f"{table}?event_id=eq.{encoded}", prefer="return=minimal")


def rows_exist(table: str, event_id: str, block_ids: list[str]) -> bool:
    if not block_ids:
        return False
    encoded_event_id = urllib.parse.quote(event_id)
    encoded_block_ids = ",".join(urllib.parse.quote(block_id) for block_id in block_ids)
    rows = supabase_request(
        "GET",
        f"{table}?select=event_id&event_id=eq.{encoded_event_id}&block_id=in.({encoded_block_ids})&limit=1",
    )
    return bool(rows)


def event_rows_exist(table: str, event_id: str) -> bool:
    encoded_event_id = urllib.parse.quote(event_id)
    rows = supabase_request(
        "GET",
        f"{table}?select=event_id&event_id=eq.{encoded_event_id}&limit=1",
    )
    return bool(rows)


def delete_block_rows(event_id: str, block_ids: list[str], *, force_replace: bool) -> None:
    if not block_ids:
        return
    if not force_replace and (
        rows_exist("scores", event_id, block_ids)
        or rows_exist("feedback", event_id, block_ids)
        or rows_exist("penalties", event_id, block_ids)
    ):
        raise RuntimeError(
            "El bloque ya tiene puntajes, feedback o penalizaciones. Usa --force-replace para reemplazarlo."
        )
    encoded_event_id = urllib.parse.quote(event_id)
    encoded_block_ids = ",".join(urllib.parse.quote(block_id) for block_id in block_ids)
    supabase_request(
        "DELETE",
        f"routines?event_id=eq.{encoded_event_id}&block_id=in.({encoded_block_ids})",
        prefer="return=minimal",
    )


def upload_to_supabase(
    app_data: dict[str, Any],
    *,
    slug: str,
    name: str,
    activate: bool,
    replace_event: bool = False,
    force_replace: bool = False,
) -> str:
    if activate:
        supabase_request(
            "PATCH",
            "events?is_active=eq.true",
            {"is_active": False},
            prefer="return=minimal",
        )

    event_rows = supabase_request(
        "POST",
        "events?on_conflict=slug",
        [
            {
                "slug": slug,
                "name": name,
                "source_name": app_data["sourceName"],
                "is_active": activate,
                "event_type": "event",
            }
        ],
        prefer="resolution=merge-duplicates,return=representation",
    )
    if not event_rows:
        raise RuntimeError("Supabase no devolvio el evento importado.")
    event_id = event_rows[0]["id"]
    imported_block_ids = [block.get("blockID") or stable_id(block["name"]) for block in app_data["blocks"]]
    if replace_event:
        if not force_replace and (
            event_rows_exist("scores", event_id)
            or event_rows_exist("feedback", event_id)
            or event_rows_exist("penalties", event_id)
        ):
            raise RuntimeError(
                "El evento ya tiene puntajes, feedback o penalizaciones. Usa --force-replace para reemplazarlo."
            )
        delete_event_rows(event_id)
    else:
        delete_block_rows(event_id, imported_block_ids, force_replace=force_replace)

    block_titles = {block["name"]: block.get("title", "") for block in app_data["blocks"]}
    block_ids = {block["name"]: block.get("blockID") or stable_id(block["name"]) for block in app_data["blocks"]}
    block_rows = [
        {
            "event_id": event_id,
            "block_id": block.get("blockID") or stable_id(block["name"]),
            "name": block["name"],
            "title": block.get("title", ""),
            "sort_order": block.get("sortOrder", index),
            "is_active": bool(block.get("isActive", index == 1)),
        }
        for index, block in enumerate(app_data["blocks"], start=1)
    ]
    routine_rows = [
        {
            "event_id": event_id,
            "routine_id": routine["id"],
            "block_id": routine.get("blockID") or block_ids.get(routine["block"], stable_id(routine["block"])),
            "block": routine["block"],
            "block_title": block_titles.get(routine["block"], ""),
            "sort_order": index,
            "name": routine["name"],
            "academy": routine["academy"],
            "division": routine["division"],
            "genre": routine["genre"],
            "level": routine["level"],
            "category": routine["category"],
            "choreographer": routine["choreographer"],
            "participant": routine.get("participant", ""),
            "state": routine["state"],
            "scheduled_time": routine["time"],
            "duration": routine["duration"],
        }
        for index, routine in enumerate(app_data["routines"], start=1)
    ]
    judge_rows = [
        {
            "event_id": event_id,
            "judge_id": stable_id(judge),
            "name": judge,
            "role": role_for_person(judge),
            "sort_order": index,
            "hero_image_name": hero_image_for_person(judge),
        }
        for index, judge in enumerate(app_data["judges"], start=1)
    ]
    template_rows = []
    criterion_rows = []
    for template_index, template in enumerate(app_data["templates"], start=1):
        template_id = stable_id(template["genre"])
        template_rows.append(
            {
                "event_id": event_id,
                "template_id": template_id,
                "genre": template["genre"],
                "title": template["title"],
                "max_score": template["maxScore"],
                "sort_order": template_index,
            }
        )
        for criterion_index, criterion in enumerate(template["criteria"], start=1):
            criterion_rows.append(
                {
                    "event_id": event_id,
                    "template_id": template_id,
                    "criterion_id": criterion["id"],
                    "section": criterion["section"],
                    "label": criterion["label"],
                    "max_score": criterion["maxScore"],
                    "sort_order": criterion_index,
                }
            )

    upsert_rows("blocks", block_rows, "event_id,block_id")
    upsert_rows("routines", routine_rows, "event_id,routine_id")
    upsert_rows("judges", judge_rows, "event_id,judge_id")
    upsert_rows("criteria_templates", template_rows, "event_id,template_id")
    upsert_rows("criteria", criterion_rows, "event_id,template_id,criterion_id")
    return event_id


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Importa el Excel de jueceo a JSON local y opcionalmente Supabase.")
    parser.add_argument("source", nargs="?", default=str(DEFAULT_XLSX), help="Ruta al .xlsx de origen.")
    parser.add_argument("output", nargs="?", default=str(DEFAULT_OUTPUT), help="Ruta del JSON de salida.")
    parser.add_argument("--supabase", action="store_true", help="Tambien sube el evento a Supabase.")
    parser.add_argument("--event-slug", default=os.environ.get("JUECEO_EVENT_SLUG", ""), help="Slug estable del evento.")
    parser.add_argument("--event-name", default=os.environ.get("JUECEO_EVENT_NAME", ""), help="Nombre visible del evento.")
    parser.add_argument("--no-activate", action="store_true", help="No marcar este evento como activo.")
    parser.add_argument("--allow-errors", action="store_true", help="Escribe/sube aunque haya errores de validacion.")
    parser.add_argument("--strict", action="store_true", help="Cancela tambien la salida JSON local si hay errores.")
    parser.add_argument("--replace-event", action="store_true", help="Reemplaza todo el evento remoto en vez de solo los bloques importados.")
    parser.add_argument("--force-replace", action="store_true", help="Permite reemplazar bloques/eventos que ya tienen puntajes o feedback.")
    parser.add_argument(
        "--template-source",
        default=os.environ.get("JUECEO_TEMPLATE_SOURCE", str(DEFAULT_TEMPLATE_XLSX)),
        help="Excel con hojas de jueceo base para programas que solo traen bloques/rutinas.",
    )
    parser.add_argument(
        "--no-template-fallback",
        action="store_true",
        help="No completar plantillas desde el Excel base cuando el origen no las incluye.",
    )
    return parser


def main() -> int:
    load_env()
    parser = build_parser()
    args = parser.parse_args()

    source = Path(args.source).expanduser().resolve()
    output = Path(args.output).expanduser().resolve()
    template_source = None if args.no_template_fallback else Path(args.template_source).expanduser().resolve()
    slug = args.event_slug or stable_id(source.stem)
    name = args.event_name or source.stem

    result = parse_workbook(source, template_source=template_source)
    for warning in result.warnings:
        print(f"WARNING: {warning}", file=sys.stderr)
    for error in result.errors:
        print(f"ERROR: {error}", file=sys.stderr)

    should_cancel = result.errors and (args.strict or (args.supabase and not args.allow_errors))
    if should_cancel:
        print("Import cancelado por errores de validacion. Usa --allow-errors solo para pruebas.", file=sys.stderr)
        return 2

    write_json(result.app_data, output)
    print(
        f"Wrote {output} with "
        f"{len(result.app_data['routines'])} routines, "
        f"{len(result.app_data['templates'])} templates, "
        f"{len(result.app_data['judges'])} judges."
    )

    if args.supabase:
        event_id = upload_to_supabase(
            result.app_data,
            slug=slug,
            name=name,
            activate=not args.no_activate,
            replace_event=args.replace_event,
            force_replace=args.force_replace,
        )
        print(f"Uploaded Supabase event '{name}' ({slug}) with id {event_id}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
